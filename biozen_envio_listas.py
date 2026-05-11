#!/usr/bin/env python3
"""
=============================================================================
  BIO ZEN — SISTEMA DE ENVÍO AUTOMÁTICO DE LISTAS DE PRECIOS
=============================================================================

Este script:
  1. Se conecta a Odoo vía XML-RPC
  2. Extrae productos y genera las 5 listas XLSX con diseño empresarial
  3. Obtiene los contactos mayoristas marcados en Odoo
  4. Envía las listas por email vía Gmail SMTP

REQUISITOS:
  pip install openpyxl python-dotenv schedule

CONFIGURACIÓN:
  Crea un archivo .env (ver env.template)

USO:
  python biozen_envio_listas.py --discover       # Ver categorías y contactos
  python biozen_envio_listas.py --generate       # Solo generar listas
  python biozen_envio_listas.py --send           # Generar + enviar ahora
  python biozen_envio_listas.py --send-to email  # Enviar a un email específico
  python biozen_envio_listas.py --schedule       # Modo programado (cron interno)
=============================================================================
"""

import xmlrpc.client
import smtplib
import os
import sys
import json
import logging
import argparse
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from pathlib import Path
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XlImage
except ImportError:
    print("ERROR: Falta openpyxl. Instálalo con: pip install openpyxl")
    sys.exit(1)

try:
    import schedule
except ImportError:
    schedule = None

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# Odoo
ODOO_URL = os.getenv('ODOO_URL', 'https://grupo-biozen130426.odoo.com')
ODOO_DB = os.getenv('ODOO_DB', 'grupo-biozen130426')
ODOO_USER = os.getenv('ODOO_USER', 'procesos@grupobiozen.com')
ODOO_PASSWORD = os.getenv('ODOO_PASSWORD', '73670480d505734b15b41a73e326221071fd8074')
ODOO_PRICELIST_ID = int(os.getenv('ODOO_PRICELIST_ID', '2'))

# Email
SMTP_SERVER = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.getenv('SMTP_PORT', '587'))
SMTP_USER = os.getenv('SMTP_USER', '')
SMTP_PASSWORD = os.getenv('SMTP_PASSWORD', '')  # App Password de Gmail
EMAIL_FROM_NAME = os.getenv('EMAIL_FROM_NAME', 'Bio Zen')

# Rutas
OUTPUT_DIR = os.getenv('OUTPUT_DIR', './listas_de_precios')
LOGO_PATH = os.getenv('LOGO_PATH', './Logo_Bio_Zen_Solo.png')

# Programación
SCHEDULE_DAY = os.getenv('SCHEDULE_DAY', 'monday')     # monday, tuesday, etc.
SCHEDULE_TIME = os.getenv('SCHEDULE_TIME', '06:00')     # HH:MM

# Campo personalizado en Odoo para filtrar mayoristas
# Opciones:
#   1. Usar una etiqueta/tag: MAYORISTA_TAG_NAME = "Mayorista"
#   2. Usar una categoría de contacto: MAYORISTA_CATEGORY = "Mayorista"
#   3. Usar un campo booleano personalizado: MAYORISTA_FIELD = "x_es_mayorista"
MAYORISTA_TAG_NAME = os.getenv('MAYORISTA_TAG_NAME', 'Mayorista')
MAYORISTA_CATEGORY = os.getenv('MAYORISTA_CATEGORY', '')
MAYORISTA_FIELD = os.getenv('MAYORISTA_FIELD', '')

# Listas a enviar por defecto (todas)
LISTAS_ENVIAR = os.getenv('LISTAS_ENVIAR', 'Naturista,Planta,Veladora,Bultos,Esoterico').split(',')


# ─────────────────────────────────────────────────────────────────────────────
# CONEXIÓN ODOO
# ─────────────────────────────────────────────────────────────────────────────

class OdooConnection:
    def __init__(self):
        self.url = ODOO_URL.rstrip('/')
        self.uid = None
        self.models = None

    def connect(self):
        logger.info(f"Conectando a Odoo: {self.url}")
        common = xmlrpc.client.ServerProxy(f'{self.url}/xmlrpc/2/common')
        self.uid = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASSWORD, {})
        if not self.uid:
            raise ConnectionError("Autenticación fallida")
        self.models = xmlrpc.client.ServerProxy(f'{self.url}/xmlrpc/2/object')
        logger.info(f"✓ Conectado como UID: {self.uid}")

    def search_read(self, model, domain, fields, limit=0, order=''):
        kwargs = {'fields': fields}
        if limit: kwargs['limit'] = limit
        if order: kwargs['order'] = order
        return self.models.execute_kw(
            ODOO_DB, self.uid, ODOO_PASSWORD,
            model, 'search_read', [domain], kwargs
        )

    def execute(self, model, method, *args, **kwargs):
        return self.models.execute_kw(
            ODOO_DB, self.uid, ODOO_PASSWORD,
            model, method, list(args), kwargs
        )


# ─────────────────────────────────────────────────────────────────────────────
# OBTENER CONTACTOS MAYORISTAS
# ─────────────────────────────────────────────────────────────────────────────

def get_wholesale_contacts(odoo):
    """
    Obtiene contactos mayoristas de Odoo.
    Filtra por tag, categoría de contacto, o campo personalizado.
    """
    domain = [('email', '!=', False), ('email', '!=', '')]

    # Método 1: Filtrar por tag/etiqueta del contacto
    if MAYORISTA_TAG_NAME:
        tags = odoo.search_read(
            'res.partner.category',
            [('name', '=', MAYORISTA_TAG_NAME)],
            ['id']
        )
        if tags:
            domain.append(('category_id', 'in', [tags[0]['id']]))
            logger.info(f"Filtrando por etiqueta: '{MAYORISTA_TAG_NAME}' (ID: {tags[0]['id']})")
        else:
            logger.warning(f"⚠ Etiqueta '{MAYORISTA_TAG_NAME}' no encontrada en Odoo")
            logger.info("  Creando etiqueta automáticamente...")
            tag_id = odoo.execute('res.partner.category', 'create', {'name': MAYORISTA_TAG_NAME})
            logger.info(f"  ✓ Etiqueta creada con ID: {tag_id}")
            logger.info("  Marca tus contactos mayoristas con esta etiqueta en Odoo")
            return []

    # Método 2: Filtrar por categoría comercial
    elif MAYORISTA_CATEGORY:
        domain.append(('category_id.name', '=', MAYORISTA_CATEGORY))

    # Método 3: Filtrar por campo personalizado booleano
    elif MAYORISTA_FIELD:
        domain.append((MAYORISTA_FIELD, '=', True))

    contacts = odoo.search_read(
        'res.partner',
        domain,
        ['id', 'name', 'email', 'phone', 'category_id', 'customer_rank',
         'active', 'property_product_pricelist', 'state_id', 'city'],
        order='name'
    )

    # Filtrar solo contactos activos con email válido
    valid = []
    for c in contacts:
        if c['active'] and c['email'] and '@' in c['email']:
            # Extraer info de la tarifa asignada
            pl = c.get('property_product_pricelist')
            c['pricelist_id'] = pl[0] if isinstance(pl, (list, tuple)) and pl else None
            c['pricelist_name'] = pl[1] if isinstance(pl, (list, tuple)) and len(pl) > 1 else 'Sin tarifa'
            valid.append(c)

    logger.info(f"✓ {len(valid)} contactos mayoristas encontrados")

    # Mostrar resumen por tarifa
    by_pricelist = defaultdict(list)
    for c in valid:
        by_pricelist[c['pricelist_name']].append(c)
    for pl_name, contacts_pl in by_pricelist.items():
        logger.info(f"  → {pl_name}: {len(contacts_pl)} contactos")

    return valid


def discover_contacts(odoo):
    """Muestra todas las etiquetas, tarifas y contactos disponibles."""

    # Mostrar listas de precios disponibles
    print("\n" + "="*70)
    print("  LISTAS DE PRECIOS (TARIFAS) DISPONIBLES")
    print("="*70)
    pricelists = odoo.search_read(
        'product.pricelist',
        [],
        ['id', 'name', 'currency_id', 'active'],
        order='name'
    )
    for pl in pricelists:
        currency = pl['currency_id'][1] if pl['currency_id'] else 'N/A'
        status = "✓ Activa" if pl['active'] else "✗ Inactiva"
        print(f"  ID: {pl['id']:<4} {pl['name']:<40} {currency:<6} {status}")

    # Mostrar etiquetas
    print(f"\n{'='*70}")
    print("  ETIQUETAS DE CONTACTO DISPONIBLES")
    print("="*70)

    categories = odoo.search_read(
        'res.partner.category',
        [],
        ['id', 'name'],
        order='name'
    )
    for cat in categories:
        count = len(odoo.search_read(
            'res.partner',
            [('category_id', 'in', [cat['id']]), ('email', '!=', False)],
            ['id']
        ))
        print(f"  ID: {cat['id']:<6} {cat['name']:<30} ({count} contactos con email)")

    # Mostrar contactos mayoristas con su tarifa
    contacts = get_wholesale_contacts(odoo)
    if contacts:
        print(f"\n{'='*70}")
        print(f"  CONTACTOS MAYORISTAS ({len(contacts)})")
        print("="*70)
        print(f"  {'Nombre':<30} {'Email':<30} {'Tarifa asignada'}")
        print(f"  {'-'*30} {'-'*30} {'-'*25}")
        for c in contacts:
            location = ''
            if c.get('city'):
                location = f" ({c['city']})"
            print(f"  {c['name']:<30} {c['email']:<30} {c['pricelist_name']}{location}")

        # Resumen por tarifa
        by_pricelist = defaultdict(list)
        for c in contacts:
            by_pricelist[c['pricelist_name']].append(c)
        print(f"\n  Resumen por tarifa:")
        for pl_name, pl_contacts in sorted(by_pricelist.items()):
            print(f"    {pl_name}: {len(pl_contacts)} contactos")

    else:
        print(f"\n⚠ No hay contactos mayoristas.")
        print(f"  Para marcar contactos:")
        print(f"  1. Ve a Contactos en Odoo")
        print(f"  2. Abre el contacto mayorista")
        print(f"  3. Agrega la etiqueta '{MAYORISTA_TAG_NAME}'")
        print(f"  4. Asegúrate de que tenga asignada su Lista de precios correcta")
        print(f"  5. Guarda")

    return contacts


# ─────────────────────────────────────────────────────────────────────────────
# GENERAR XLSX (con diseño empresarial Bio Zen)
# ─────────────────────────────────────────────────────────────────────────────

LISTA_CONFIG = {
    "Naturista": {
        "filename": "Lista_de_precio_Mayorista_Naturista.xlsx",
        "display": "NATURISTA",
        "categ_names": ["Naturista", "Natural", "Herbolaria"],
    },
    "Planta": {
        "filename": "Lista_de_precio_Mayorista_Planta.xlsx",
        "display": "PLANTAS Y HIERBAS",
        "categ_names": ["Planta", "Hierba"],
    },
    "Veladora": {
        "filename": "Lista_de_precio_Mayorista_Veladora.xlsx",
        "display": "VELADORAS",
        "categ_names": ["Veladora", "Vela"],
    },
    "Bultos": {
        "filename": "Lista_de_precio_Mayorista_Bultos.xlsx",
        "display": "BULTOS Y FIGURAS",
        "categ_names": ["Bulto", "Figura"],
    },
    "Esoterico": {
        "filename": "Lista_de_precio_Mayorista_Esoterico.xlsx",
        "display": "ESOTÉRICO",
        "categ_names": ["Esotérico", "Esoterico", "Incienso"],
    },
}

# Colors
BLACK = '1A1A2E'
DARK = '16213E'
ACCENT = '0F3460'
GOLD = 'C9A227'
LIGHT_GRAY = 'F5F5F5'
BRAND_BG = 'E8EEF2'
BRAND_DARK = '2C3E50'


def get_products_for_list(odoo, config, pricelist_id=None):
    """Extrae productos filtrados desde Odoo con precios de la tarifa indicada."""
    if pricelist_id is None:
        pricelist_id = ODOO_PRICELIST_ID

    domain = [('sale_ok', '=', True), ('active', '=', True)]

    if config.get('categ_names'):
        cat_ids = []
        for name in config['categ_names']:
            cats = odoo.search_read('product.category', [('complete_name', 'ilike', name)], ['id'])
            cat_ids.extend([c['id'] for c in cats])
        if cat_ids:
            domain.append(('categ_id', 'in', cat_ids))

    if config.get('custom_domain'):
        domain.extend(config['custom_domain'])

    fields = ['default_code', 'name', 'list_price', 'categ_id']

    # Try to get brand field
    brand_field = None
    try:
        fields_info = odoo.execute('product.template', 'fields_get', [], attributes=['string'])
        for f in ['product_brand_id', 'brand_id', 'x_brand', 'x_marca']:
            if f in fields_info:
                brand_field = f
                fields.append(f)
                break
    except Exception:
        pass

    products = odoo.search_read('product.template', domain, fields, order='name')

    result = []
    for prod in products:
        try:
            price = odoo.execute('product.pricelist', 'get_product_price_rule',
                                 [pricelist_id], prod['id'], 1.0, False)
            mayoreo_price = price[0] if isinstance(price, (list, tuple)) else price
        except Exception:
            mayoreo_price = prod.get('list_price', 0)

        brand = ''
        if brand_field and prod.get(brand_field):
            val = prod[brand_field]
            brand = val[1] if isinstance(val, (list, tuple)) else str(val)

        result.append({
            'ref': prod.get('default_code', ''),
            'name': prod.get('name', ''),
            'price': mayoreo_price,
            'brand': brand.upper() if brand else 'SIN MARCA',
        })

    return result


def generate_xlsx(list_name, products, config, pricelist_name=None, output_subdir=None):
    """Genera XLSX con diseño empresarial Bio Zen."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Precios"

    # Page setup
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 65
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 3

    # Styles
    thin_border = Border(bottom=Side(style='hair', color='D0D0D0'))
    gold_border = Border(bottom=Side(style='medium', color=GOLD))
    brand_border = Border(
        top=Side(style='thin', color=ACCENT),
        bottom=Side(style='thin', color=ACCENT)
    )

    # Header section (rows 1-8)
    for r in range(1, 9):
        ws.row_dimensions[r].height = 18
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=BLACK)

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[7].height = 6
    ws.row_dimensions[8].height = 4

    # Logo
    if os.path.exists(LOGO_PATH):
        logo = XlImage(LOGO_PATH)
        logo.width = 70
        logo.height = 78
        ws.add_image(logo, 'B2')

    # Company name
    ws.merge_cells('C2:D2')
    c = ws.cell(row=2, column=3, value='BIO ZEN')
    c.font = Font(name='Arial', bold=True, size=22, color=GOLD)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill = PatternFill('solid', fgColor=BLACK)

    # Subtitle
    ws.merge_cells('C3:D3')
    c = ws.cell(row=3, column=3, value='DISTRIBUIDORA MAYORISTA')
    c.font = Font(name='Arial', size=11, color='AAAAAA')
    c.alignment = Alignment(horizontal='left', vertical='top')
    c.fill = PatternFill('solid', fgColor=BLACK)

    # List name + tarifa
    ws.merge_cells('C4:D4')
    display_name = config.get('display', list_name.upper())
    tarifa_text = f' · Tarifa: {pricelist_name}' if pricelist_name else ''
    c = ws.cell(row=4, column=3, value=f'Lista de Precios Mayoreo — {display_name}{tarifa_text}')
    c.font = Font(name='Arial', size=10, color=GOLD, italic=True)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill = PatternFill('solid', fgColor=BLACK)

    # Date + counts
    brands = defaultdict(list)
    for p in products:
        brands[p['brand']].append(p)
    brand_count = len(brands)

    today = datetime.now().strftime('%d/%m/%Y')
    ws.merge_cells('C5:D5')
    c = ws.cell(row=5, column=3, value=f'Fecha: {today}  ·  {len(products)} productos  ·  {brand_count} marcas')
    c.font = Font(name='Arial', size=8.5, color='888888')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill = PatternFill('solid', fgColor=BLACK)

    # Gold line
    for col in range(2, 5):
        ws.cell(row=6, column=col).fill = PatternFill('solid', fgColor=BLACK)
        ws.cell(row=6, column=col).border = gold_border

    # White spacer
    for col in range(1, 6):
        ws.cell(row=8, column=col).fill = PatternFill('solid', fgColor='FFFFFF')

    # Column headers (row 9)
    row = 9
    ws.row_dimensions[row].height = 28
    for col, title in [(2, 'Referencia'), (3, 'Producto'), (4, 'Precio Mayoreo')]:
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=DARK)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = gold_border

    # Data rows
    row = 10
    sorted_brands = sorted(brands.keys())
    alt = False

    for brand in sorted_brands:
        brand_prods = sorted(brands[brand], key=lambda x: x['name'])

        # Brand header
        ws.row_dimensions[row].height = 22
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=2, value=f"{brand} ({len(brand_prods)})")
        cell.font = Font(name='Arial', bold=True, size=10, color=BRAND_DARK)
        cell.fill = PatternFill('solid', fgColor=BRAND_BG)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = brand_border
        for c in [3, 4]:
            ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=BRAND_BG)
            ws.cell(row=row, column=c).border = brand_border
        row += 1
        alt = False

        for p in brand_prods:
            ws.row_dimensions[row].height = 17
            fill = PatternFill('solid', fgColor=LIGHT_GRAY) if alt else PatternFill('solid', fgColor='FFFFFF')

            ws.cell(row=row, column=2, value=p['ref']).font = Font(name='Arial', size=9, color='666666')
            ws.cell(row=row, column=2).fill = fill
            ws.cell(row=row, column=2).border = thin_border

            ws.cell(row=row, column=3, value=p['name']).font = Font(name='Arial', size=9.5, color='333333')
            ws.cell(row=row, column=3).fill = fill
            ws.cell(row=row, column=3).border = thin_border

            price_cell = ws.cell(row=row, column=4, value=p['price'])
            price_cell.font = Font(name='Arial', bold=True, size=10, color=ACCENT)
            price_cell.fill = fill
            price_cell.number_format = '$#,##0.00'
            price_cell.alignment = Alignment(horizontal='right')
            price_cell.border = thin_border

            row += 1
            alt = not alt

    # Footer
    row += 1
    for c in range(2, 5):
        ws.cell(row=row, column=c).border = Border(top=Side(style='medium', color=GOLD))

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2, value='Precios sujetos a cambio sin previo aviso · Los precios no incluyen IVA')
    c.font = Font(name='Arial', size=8, color='999999', italic=True)
    c.alignment = Alignment(horizontal='center')

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2, value=f'Bio Zen · Lista generada el {today}')
    c.font = Font(name='Arial', size=8, color='AAAAAA')
    c.alignment = Alignment(horizontal='center')

    ws.freeze_panes = 'B10'
    ws.print_title_rows = '1:9'

    # Save — optionally in a pricelist-specific subdirectory
    if output_subdir:
        output_path = Path(OUTPUT_DIR) / output_subdir
    else:
        output_path = Path(OUTPUT_DIR)
    output_path.mkdir(parents=True, exist_ok=True)
    filepath = output_path / config['filename']
    wb.save(str(filepath))
    logger.info(f"✓ {list_name}: {len(products)} productos → {filepath}")
    return str(filepath)


# ─────────────────────────────────────────────────────────────────────────────
# ENVÍO DE EMAIL
# ─────────────────────────────────────────────────────────────────────────────

def create_email_html(contact_name, lists_info, pricelist_name=None):
    """Genera el cuerpo HTML del email."""
    today = datetime.now().strftime('%d de %B, %Y')

    tarifa_line = ''
    if pricelist_name:
        tarifa_line = f'<p style="font-size:13px;color:#0F3460;background:#E8EEF2;padding:10px 16px;border-radius:8px;margin:16px 0;">Tarifa aplicada: <strong>{pricelist_name}</strong></p>'

    lists_html = ""
    for info in lists_info:
        lists_html += f"""
        <tr>
          <td style="padding:8px 16px;border-bottom:1px solid #eee;">
            <span style="font-size:18px;margin-right:8px;">{info['icon']}</span>
            <strong>{info['name']}</strong>
          </td>
          <td style="padding:8px 16px;border-bottom:1px solid #eee;text-align:right;color:#666;">
            {info['count']} productos
          </td>
        </tr>"""

    return f"""
    <html>
    <body style="font-family:Arial,sans-serif;color:#333;margin:0;padding:0;background:#f5f5f5;">
      <div style="max-width:600px;margin:20px auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.1);">

        <!-- Header -->
        <div style="background:#1A1A2E;padding:30px;text-align:center;">
          <h1 style="color:#C9A227;margin:0;font-size:28px;letter-spacing:2px;">BIO ZEN</h1>
          <p style="color:#888;margin:8px 0 0;font-size:13px;">DISTRIBUIDORA MAYORISTA</p>
        </div>

        <!-- Body -->
        <div style="padding:30px;">
          <p style="font-size:15px;line-height:1.6;">
            Estimado/a <strong>{contact_name}</strong>,
          </p>
          <p style="font-size:14px;line-height:1.6;color:#555;">
            Le compartimos nuestras listas de precios mayorista actualizadas al <strong>{today}</strong>.
            Encontrará adjuntos los siguientes catálogos:
          </p>

          {tarifa_line}

          <table style="width:100%;border-collapse:collapse;margin:20px 0;border-radius:8px;overflow:hidden;border:1px solid #eee;">
            <thead>
              <tr style="background:#f8f8f8;">
                <th style="padding:10px 16px;text-align:left;font-size:13px;color:#666;">Lista</th>
                <th style="padding:10px 16px;text-align:right;font-size:13px;color:#666;">Productos</th>
              </tr>
            </thead>
            <tbody>{lists_html}</tbody>
          </table>

          <p style="font-size:14px;line-height:1.6;color:#555;">
            Los precios están sujetos a cambio sin previo aviso y no incluyen IVA.
            Para pedidos o dudas, no dude en contactarnos.
          </p>

          <p style="font-size:14px;margin-top:24px;">
            Saludos cordiales,<br>
            <strong style="color:#1A1A2E;">Equipo Bio Zen</strong>
          </p>
        </div>

        <!-- Footer -->
        <div style="background:#f8f8f8;padding:16px 30px;text-align:center;border-top:1px solid #eee;">
          <p style="font-size:11px;color:#999;margin:0;">
            Bio Zen · Distribuidora Mayorista<br>
            Este correo fue generado automáticamente
          </p>
        </div>

      </div>
    </body>
    </html>
    """


def send_email(to_email, to_name, subject, html_body, attachments):
    """Envía un email con archivos adjuntos vía SMTP."""
    if not SMTP_USER or not SMTP_PASSWORD:
        logger.error("⚠ Configura SMTP_USER y SMTP_PASSWORD en el .env")
        return False

    msg = MIMEMultipart()
    msg['From'] = f'{EMAIL_FROM_NAME} <{SMTP_USER}>'
    msg['To'] = to_email
    msg['Subject'] = subject

    # HTML body
    msg.attach(MIMEText(html_body, 'html'))

    # Attachments
    for filepath in attachments:
        if not os.path.exists(filepath):
            logger.warning(f"  ⚠ Archivo no encontrado: {filepath}")
            continue

        with open(filepath, 'rb') as f:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            filename = os.path.basename(filepath)
            part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
            msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
        logger.info(f"  ✓ Email enviado a: {to_name} <{to_email}>")
        return True
    except Exception as e:
        logger.error(f"  ✗ Error enviando a {to_email}: {e}")
        return False


def send_to_contacts(odoo, generated_files):
    """Envía las listas a todos los contactos mayoristas."""
    contacts = get_wholesale_contacts(odoo)

    if not contacts:
        logger.warning("⚠ No hay contactos mayoristas para enviar")
        return

    today = datetime.now().strftime('%d/%m/%Y')
    subject = f'Bio Zen — Listas de Precios Mayorista ({today})'

    # Prepare lists info for email
    icons = {'Naturista': '🌿', 'Planta': '🌱', 'Veladora': '🕯️', 'Bultos': '🗿', 'Esoterico': '🔮'}
    lists_info = []
    attachments = []
    for name, filepath in generated_files.items():
        if os.path.exists(filepath):
            attachments.append(filepath)
            # Count products (approximate from filename)
            lists_info.append({
                'name': LISTA_CONFIG[name]['display'],
                'icon': icons.get(name, '📋'),
                'count': '—',
            })

    sent = 0
    failed = 0
    for contact in contacts:
        html = create_email_html(contact['name'], lists_info)
        if send_email(contact['email'], contact['name'], subject, html, attachments):
            sent += 1
        else:
            failed += 1
        # Small delay to avoid rate limiting
        time.sleep(1)

    logger.info(f"\n✓ Resumen: {sent} enviados, {failed} fallidos de {len(contacts)} contactos")


def send_to_single(email, generated_files):
    """Envía las listas a un email específico."""
    today = datetime.now().strftime('%d/%m/%Y')
    subject = f'Bio Zen — Listas de Precios Mayorista ({today})'

    icons = {'Naturista': '🌿', 'Planta': '🌱', 'Veladora': '🕯️', 'Bultos': '🗿', 'Esoterico': '🔮'}
    lists_info = []
    attachments = []
    for name, filepath in generated_files.items():
        if os.path.exists(filepath):
            attachments.append(filepath)
            lists_info.append({
                'name': LISTA_CONFIG[name]['display'],
                'icon': icons.get(name, '📋'),
                'count': '—',
            })

    html = create_email_html(email.split('@')[0], lists_info)
    send_email(email, email, subject, html, attachments)


# ─────────────────────────────────────────────────────────────────────────────
# FLUJO PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

def generate_all_lists(odoo, pricelist_id=None, pricelist_name=None, output_subdir=None):
    """Genera todas las listas XLSX para una tarifa específica."""
    generated = {}
    for list_name, config in LISTA_CONFIG.items():
        if list_name not in LISTAS_ENVIAR:
            continue
        products = get_products_for_list(odoo, config, pricelist_id)
        if products:
            filepath = generate_xlsx(list_name, products, config, pricelist_name, output_subdir)
            generated[list_name] = filepath
        else:
            logger.warning(f"⚠ Sin productos para '{list_name}'")
    return generated


def generate_lists_per_pricelist(odoo, contacts):
    """
    Genera listas agrupadas por tarifa.
    Retorna: { pricelist_id: { 'name': str, 'files': {list_name: filepath}, 'contacts': [...] } }
    """
    by_pricelist = defaultdict(lambda: {'name': '', 'files': {}, 'contacts': []})

    for c in contacts:
        pl_id = c['pricelist_id'] or ODOO_PRICELIST_ID
        pl_name = c['pricelist_name']
        by_pricelist[pl_id]['name'] = pl_name
        by_pricelist[pl_id]['contacts'].append(c)

    for pl_id, data in by_pricelist.items():
        pl_name = data['name']
        # Crear subdirectorio con nombre de tarifa limpio
        safe_name = ''.join(c if c.isalnum() or c in ' _-' else '_' for c in pl_name).strip()
        subdir = safe_name or f'tarifa_{pl_id}'

        logger.info(f"\n{'─'*50}")
        logger.info(f"Generando listas para tarifa: {pl_name} (ID: {pl_id})")
        logger.info(f"  → {len(data['contacts'])} contactos con esta tarifa")
        logger.info(f"{'─'*50}")

        data['files'] = generate_all_lists(odoo, pl_id, pl_name, subdir)

    return by_pricelist


def send_to_contacts(odoo, contacts=None):
    """Envía las listas personalizadas por tarifa a todos los contactos mayoristas."""
    if contacts is None:
        contacts = get_wholesale_contacts(odoo)

    if not contacts:
        logger.warning("⚠ No hay contactos mayoristas para enviar")
        return

    # Generar listas agrupadas por tarifa
    pricelist_data = generate_lists_per_pricelist(odoo, contacts)

    today = datetime.now().strftime('%d/%m/%Y')
    icons = {'Naturista': '🌿', 'Planta': '🌱', 'Veladora': '🕯️', 'Bultos': '🗿', 'Esoterico': '🔮'}

    sent = 0
    failed = 0

    for pl_id, data in pricelist_data.items():
        if not data['files']:
            logger.warning(f"⚠ Sin archivos para tarifa '{data['name']}', saltando")
            continue

        # Preparar info de las listas para el email
        lists_info = []
        attachments = []
        for name, filepath in data['files'].items():
            if os.path.exists(filepath):
                attachments.append(filepath)
                lists_info.append({
                    'name': LISTA_CONFIG[name]['display'],
                    'icon': icons.get(name, '📋'),
                    'count': '—',
                })

        subject = f'Bio Zen — Listas de Precios Mayorista ({today})'

        # Enviar a cada contacto de esta tarifa
        for contact in data['contacts']:
            logger.info(f"  Enviando a {contact['name']} (tarifa: {data['name']})")
            html = create_email_html(contact['name'], lists_info, data['name'])
            if send_email(contact['email'], contact['name'], subject, html, attachments):
                sent += 1
            else:
                failed += 1
            time.sleep(1)

    logger.info(f"\n{'='*60}")
    logger.info(f"✓ Resumen: {sent} enviados, {failed} fallidos de {len(contacts)} contactos")
    logger.info(f"  Tarifas procesadas: {len(pricelist_data)}")
    for pl_id, data in pricelist_data.items():
        logger.info(f"    {data['name']}: {len(data['contacts'])} contactos, {len(data['files'])} listas")
    logger.info(f"{'='*60}")


def full_run():
    """Ejecuta el ciclo completo: generar + enviar."""
    logger.info("═" * 60)
    logger.info("  BIO ZEN — Envío automático de listas de precios")
    logger.info(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("═" * 60)

    odoo = OdooConnection()
    odoo.connect()

    contacts = get_wholesale_contacts(odoo)
    if contacts:
        send_to_contacts(odoo, contacts)
    else:
        logger.error("No hay contactos mayoristas")


def main():
    parser = argparse.ArgumentParser(description='Bio Zen - Envío de listas de precios')
    parser.add_argument('--discover', action='store_true', help='Ver categorías y contactos')
    parser.add_argument('--generate', action='store_true', help='Solo generar listas')
    parser.add_argument('--send', action='store_true', help='Generar y enviar a mayoristas')
    parser.add_argument('--send-to', type=str, help='Enviar a un email específico')
    parser.add_argument('--schedule', action='store_true', help='Modo programado')
    args = parser.parse_args()

    if not any([args.discover, args.generate, args.send, args.send_to, args.schedule]):
        parser.print_help()
        return

    # Check config
    if 'tu-instancia' in ODOO_URL or 'tu_contraseña' in ODOO_PASSWORD:
        print("\n⚠️  Configura el archivo .env con tus credenciales de Odoo")
        return

    odoo = OdooConnection()
    odoo.connect()

    if args.discover:
        discover_contacts(odoo)
        return

    if args.generate:
        # Generar con la tarifa por defecto o todas las tarifas de los contactos
        contacts = get_wholesale_contacts(odoo)
        if contacts:
            generate_lists_per_pricelist(odoo, contacts)
        else:
            generate_all_lists(odoo)
        return

    if args.send_to:
        # Para envío individual, usar la tarifa por defecto
        generated = generate_all_lists(odoo)
        if generated:
            send_to_single(args.send_to, generated)
        return

    if args.send:
        contacts = get_wholesale_contacts(odoo)
        if contacts:
            send_to_contacts(odoo, contacts)
        return

    if args.schedule:
        if not schedule:
            print("ERROR: Instala schedule: pip install schedule")
            return

        day_map = {
            'monday': schedule.every().monday,
            'tuesday': schedule.every().tuesday,
            'wednesday': schedule.every().wednesday,
            'thursday': schedule.every().thursday,
            'friday': schedule.every().friday,
            'saturday': schedule.every().saturday,
            'sunday': schedule.every().sunday,
        }

        job = day_map.get(SCHEDULE_DAY.lower(), schedule.every().monday)
        job.at(SCHEDULE_TIME).do(full_run)

        logger.info(f"Programado: cada {SCHEDULE_DAY} a las {SCHEDULE_TIME}")
        logger.info("Esperando... (Ctrl+C para salir)")

        while True:
            schedule.run_pending()
            time.sleep(60)


if __name__ == '__main__':
    main()
