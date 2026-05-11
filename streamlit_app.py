import streamlit as st
import xmlrpc.client
import smtplib
import os
import io
import time
import tempfile
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Bio Zen — Listas de Precios",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────────────────────────────────────

st.markdown("""
<style>
    .main .block-container { padding-top: 2rem; max-width: 1100px; }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] {
        padding: 8px 20px; border-radius: 10px;
        font-weight: 500; font-size: 14px;
    }
    div[data-testid="stMetric"] {
        background: #f8f9fa; border-radius: 12px;
        padding: 16px; border: 1px solid #e9ecef;
    }
    .success-box {
        background: #d4edda; border: 1px solid #c3e6cb;
        border-radius: 10px; padding: 16px; margin: 8px 0;
    }
    .info-box {
        background: #e8eef2; border: 1px solid #b8d4e3;
        border-radius: 10px; padding: 16px; margin: 8px 0;
    }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

LISTA_CONFIG = {
    "Naturista": {"filename": "Lista_de_precio_Mayorista_Naturista.xlsx", "display": "NATURISTA", "icon": "🌿", "categ_names": ["Naturista", "Natural", "Herbolaria"]},
    "Planta": {"filename": "Lista_de_precio_Mayorista_Planta.xlsx", "display": "PLANTAS Y HIERBAS", "icon": "🌱", "categ_names": ["Planta", "Hierba"]},
    "Veladora": {"filename": "Lista_de_precio_Mayorista_Veladora.xlsx", "display": "VELADORAS", "icon": "🕯️", "categ_names": ["Veladora", "Vela"]},
    "Bultos": {"filename": "Lista_de_precio_Mayorista_Bultos.xlsx", "display": "BULTOS Y FIGURAS", "icon": "🗿", "categ_names": ["Bulto", "Figura"]},
    "Esoterico": {"filename": "Lista_de_precio_Mayorista_Esoterico.xlsx", "display": "ESOTÉRICO", "icon": "🔮", "categ_names": ["Esotérico", "Esoterico", "Incienso"]},
}

BLACK = '1A1A2E'
DARK = '16213E'
ACCENT = '0F3460'
GOLD = 'C9A227'
LIGHT_GRAY = 'F5F5F5'
BRAND_BG = 'E8EEF2'
BRAND_DARK = '2C3E50'

# ─────────────────────────────────────────────────────────────────────────────
# ODOO CONNECTION
# ─────────────────────────────────────────────────────────────────────────────

class OdooConnection:
    def __init__(self, url, db, user, password):
        self.url = url.rstrip('/')
        self.db = db
        self.user = user
        self.password = password
        self.uid = None
        self.models = None

    def connect(self):
        common = xmlrpc.client.ServerProxy(f'{self.url}/xmlrpc/2/common')
        self.uid = common.authenticate(self.db, self.user, self.password, {})
        if not self.uid:
            raise ConnectionError("Autenticación fallida. Verifica credenciales.")
        self.models = xmlrpc.client.ServerProxy(f'{self.url}/xmlrpc/2/object')
        return True

    def search_read(self, model, domain, fields, limit=0, order=''):
        kwargs = {'fields': fields}
        if limit: kwargs['limit'] = limit
        if order: kwargs['order'] = order
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, 'search_read', [domain], kwargs
        )

    def execute(self, model, method, *args, **kwargs):
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, method, list(args), kwargs
        )


@st.cache_resource(ttl=300)
def get_odoo_connection(url, db, user, password):
    """Cached Odoo connection."""
    odoo = OdooConnection(url, db, user, password)
    odoo.connect()
    return odoo


# ─────────────────────────────────────────────────────────────────────────────
# CONTACTS
# ─────────────────────────────────────────────────────────────────────────────

def get_wholesale_contacts(odoo, tag_name="Mayorista"):
    domain = [('email', '!=', False), ('email', '!=', '')]

    tags = odoo.search_read('res.partner.category', [('name', '=', tag_name)], ['id'])
    if tags:
        domain.append(('category_id', 'in', [tags[0]['id']]))
    else:
        return []

    contacts = odoo.search_read(
        'res.partner', domain,
        ['id', 'name', 'email', 'phone', 'active',
         'property_product_pricelist', 'state_id', 'city'],
        order='name'
    )

    valid = []
    for c in contacts:
        if c['active'] and c['email'] and '@' in c['email']:
            pl = c.get('property_product_pricelist')
            c['pricelist_id'] = pl[0] if isinstance(pl, (list, tuple)) and pl else None
            c['pricelist_name'] = pl[1] if isinstance(pl, (list, tuple)) and len(pl) > 1 else 'Sin tarifa'
            valid.append(c)

    return valid


def get_pricelists(odoo):
    return odoo.search_read(
        'product.pricelist', [],
        ['id', 'name', 'currency_id', 'active'],
        order='name'
    )


# ─────────────────────────────────────────────────────────────────────────────
# PRODUCT EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def get_products_for_list(odoo, config, pricelist_id):
    domain = [('sale_ok', '=', True), ('active', '=', True)]

    if config.get('categ_names'):
        cat_ids = []
        for name in config['categ_names']:
            cats = odoo.search_read('product.category', [('complete_name', 'ilike', name)], ['id'])
            cat_ids.extend([c['id'] for c in cats])
        if cat_ids:
            domain.append(('categ_id', 'in', cat_ids))

    fields = ['default_code', 'name', 'list_price', 'categ_id']

    brand_field = None
    try:
        fields_info = odoo.execute('product.template', 'fields_get', [], attributes=['string'])
        for f in ['product_brand_id', 'brand_id', 'x_brand', 'x_marca']:
            if f in fields_info:
                brand_field = f
                fields.append(f)
                break
    except Exception:
        pass

    products = odoo.search_read('product.template', domain, fields, order='name')

    result = []
    for prod in products:
        try:
            price = odoo.execute('product.pricelist', 'get_product_price_rule',
                                 [pricelist_id], prod['id'], 1.0, False)
            mayoreo_price = price[0] if isinstance(price, (list, tuple)) else price
        except Exception:
            mayoreo_price = prod.get('list_price', 0)

        brand = ''
        if brand_field and prod.get(brand_field):
            val = prod[brand_field]
            brand = val[1] if isinstance(val, (list, tuple)) else str(val)

        result.append({
            'ref': prod.get('default_code', ''),
            'name': prod.get('name', ''),
            'price': mayoreo_price,
            'brand': brand.upper() if brand else 'SIN MARCA',
        })

    return result


# ─────────────────────────────────────────────────────────────────────────────
# XLSX GENERATION
# ─────────────────────────────────────────────────────────────────────────────

def generate_xlsx_bytes(list_name, products, config, pricelist_name=None, logo_bytes=None):
    """Genera un XLSX en memoria y retorna los bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Precios"

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 65
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 3

    thin_border = Border(bottom=Side(style='hair', color='D0D0D0'))
    gold_border = Border(bottom=Side(style='medium', color=GOLD))
    brand_border = Border(top=Side(style='thin', color=ACCENT), bottom=Side(style='thin', color=ACCENT))

    for r in range(1, 9):
        ws.row_dimensions[r].height = 18
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=BLACK)

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[7].height = 6
    ws.row_dimensions[8].height = 4

    # Logo
    if logo_bytes:
        try:
            from openpyxl.drawing.image import Image as XlImage
            tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            tmp.write(logo_bytes)
            tmp.close()
            logo = XlImage(tmp.name)
            logo.width = 70
            logo.height = 78
            ws.add_image(logo, 'B2')
        except Exception:
            pass

    ws.merge_cells('C2:D2')
    c = ws.cell(row=2, column=3, value='BIO ZEN')
    c.font = Font(name='Arial', bold=True, size=22, color=GOLD)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill = PatternFill('solid', fgColor=BLACK)

    ws.merge_cells('C3:D3')
    c = ws.cell(row=3, column=3, value='DISTRIBUIDORA MAYORISTA')
    c.font = Font(name='Arial', size=11, color='AAAAAA')
    c.alignment = Alignment(horizontal='left', vertical='top')
    c.fill = PatternFill('solid', fgColor=BLACK)

    ws.merge_cells('C4:D4')
    display_name = config.get('display', list_name.upper())
    tarifa_text = f' · Tarifa: {pricelist_name}' if pricelist_name else ''
    c = ws.cell(row=4, column=3, value=f'Lista de Precios Mayoreo — {display_name}{tarifa_text}')
    c.font = Font(name='Arial', size=10, color=GOLD, italic=True)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill = PatternFill('solid', fgColor=BLACK)

    brands = defaultdict(list)
    for p in products:
        brands[p['brand']].append(p)
    brand_count = len(brands)
    today = datetime.now().strftime('%d/%m/%Y')

    ws.merge_cells('C5:D5')
    c = ws.cell(row=5, column=3, value=f'Fecha: {today}  ·  {len(products)} productos  ·  {brand_count} marcas')
    c.font = Font(name='Arial', size=8.5, color='888888')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill = PatternFill('solid', fgColor=BLACK)

    for col in range(2, 5):
        ws.cell(row=6, column=col).fill = PatternFill('solid', fgColor=BLACK)
        ws.cell(row=6, column=col).border = gold_border

    for col in range(1, 6):
        ws.cell(row=8, column=col).fill = PatternFill('solid', fgColor='FFFFFF')

    row = 9
    ws.row_dimensions[row].height = 28
    for col, title in [(2, 'Referencia'), (3, 'Producto'), (4, 'Precio Mayoreo')]:
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=DARK)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = gold_border

    row = 10
    sorted_brands = sorted(brands.keys())
    alt = False

    for brand in sorted_brands:
        brand_prods = sorted(brands[brand], key=lambda x: x['name'])
        ws.row_dimensions[row].height = 22
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=2, value=f"{brand} ({len(brand_prods)})")
        cell.font = Font(name='Arial', bold=True, size=10, color=BRAND_DARK)
        cell.fill = PatternFill('solid', fgColor=BRAND_BG)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = brand_border
        for c in [3, 4]:
            ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=BRAND_BG)
            ws.cell(row=row, column=c).border = brand_border
        row += 1
        alt = False

        for p in brand_prods:
            ws.row_dimensions[row].height = 17
            fill = PatternFill('solid', fgColor=LIGHT_GRAY) if alt else PatternFill('solid', fgColor='FFFFFF')

            ws.cell(row=row, column=2, value=p['ref']).font = Font(name='Arial', size=9, color='666666')
            ws.cell(row=row, column=2).fill = fill
            ws.cell(row=row, column=2).border = thin_border

            ws.cell(row=row, column=3, value=p['name']).font = Font(name='Arial', size=9.5, color='333333')
            ws.cell(row=row, column=3).fill = fill
            ws.cell(row=row, column=3).border = thin_border

            price_cell = ws.cell(row=row, column=4, value=p['price'])
            price_cell.font = Font(name='Arial', bold=True, size=10, color=ACCENT)
            price_cell.fill = fill
            price_cell.number_format = '$#,##0.00'
            price_cell.alignment = Alignment(horizontal='right')
            price_cell.border = thin_border

            row += 1
            alt = not alt

    row += 1
    for c in range(2, 5):
        ws.cell(row=row, column=c).border = Border(top=Side(style='medium', color=GOLD))

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2, value='Precios sujetos a cambio sin previo aviso · Los precios no incluyen IVA')
    c.font = Font(name='Arial', size=8, color='999999', italic=True)
    c.alignment = Alignment(horizontal='center')

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2, value=f'Bio Zen · Lista generada el {today}')
    c.font = Font(name='Arial', size=8, color='AAAAAA')
    c.alignment = Alignment(horizontal='center')

    ws.freeze_panes = 'B10'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# EMAIL
# ─────────────────────────────────────────────────────────────────────────────

def create_email_html(contact_name, lists_info, pricelist_name=None):
    today = datetime.now().strftime('%d de %B, %Y')
    tarifa_line = ''
    if pricelist_name:
        tarifa_line = f'<p style="font-size:13px;color:#0F3460;background:#E8EEF2;padding:10px 16px;border-radius:8px;margin:16px 0;">Tarifa aplicada: <strong>{pricelist_name}</strong></p>'

    lists_html = ""
    for info in lists_info:
        lists_html += f"""<tr>
          <td style="padding:8px 16px;border-bottom:1px solid #eee;"><span style="font-size:18px;margin-right:8px;">{info['icon']}</span><strong>{info['name']}</strong></td>
          <td style="padding:8px 16px;border-bottom:1px solid #eee;text-align:right;color:#666;">{info['count']} productos</td>
        </tr>"""

    return f"""<html><body style="font-family:Arial,sans-serif;color:#333;margin:0;padding:0;background:#f5f5f5;">
      <div style="max-width:600px;margin:20px auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.1);">
        <div style="background:#1A1A2E;padding:30px;text-align:center;">
          <h1 style="color:#C9A227;margin:0;font-size:28px;letter-spacing:2px;">BIO ZEN</h1>
          <p style="color:#888;margin:8px 0 0;font-size:13px;">DISTRIBUIDORA MAYORISTA</p>
        </div>
        <div style="padding:30px;">
          <p style="font-size:15px;line-height:1.6;">Estimado/a <strong>{contact_name}</strong>,</p>
          <p style="font-size:14px;line-height:1.6;color:#555;">Le compartimos nuestras listas de precios mayorista actualizadas al <strong>{today}</strong>.</p>
          {tarifa_line}
          <table style="width:100%;border-collapse:collapse;margin:20px 0;border:1px solid #eee;">
            <thead><tr style="background:#f8f8f8;">
              <th style="padding:10px 16px;text-align:left;font-size:13px;color:#666;">Lista</th>
              <th style="padding:10px 16px;text-align:right;font-size:13px;color:#666;">Productos</th>
            </tr></thead>
            <tbody>{lists_html}</tbody>
          </table>
          <p style="font-size:14px;line-height:1.6;color:#555;">Los precios están sujetos a cambio sin previo aviso y no incluyen IVA.</p>
          <p style="font-size:14px;margin-top:24px;">Saludos cordiales,<br><strong style="color:#1A1A2E;">Equipo Bio Zen</strong></p>
        </div>
        <div style="background:#f8f8f8;padding:16px 30px;text-align:center;border-top:1px solid #eee;">
          <p style="font-size:11px;color:#999;margin:0;">Bio Zen · Distribuidora Mayorista<br>Correo generado automáticamente</p>
        </div>
      </div></body></html>"""


def send_email(smtp_server, smtp_port, smtp_user, smtp_pass, from_name,
               to_email, to_name, subject, html_body, attachment_files):
    msg = MIMEMultipart()
    msg['From'] = f'{from_name} <{smtp_user}>'
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(html_body, 'html'))

    for filename, file_bytes in attachment_files:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(file_bytes)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.send_message(msg)

    return True


# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────────────────────────────────────

# Header
st.markdown("""
<div style="background:linear-gradient(135deg,#1A1A2E,#16213E);padding:24px 32px;border-radius:16px;margin-bottom:24px;">
  <h1 style="color:#C9A227;margin:0;font-size:28px;letter-spacing:1px;">📋 Bio Zen</h1>
  <p style="color:#888;margin:6px 0 0;font-size:14px;">Generador y envío automático de listas de precios mayorista</p>
</div>
""", unsafe_allow_html=True)

# ─── SIDEBAR: Configuración ───
with st.sidebar:
    st.header("⚙️ Configuración")

    st.subheader("🏢 Odoo")
    odoo_url = st.text_input("URL de Odoo", value=st.secrets.get("ODOO_URL", ""), placeholder="https://tu-instancia.odoo.com")
    odoo_db = st.text_input("Base de datos", value=st.secrets.get("ODOO_DB", ""), placeholder="nombre_db")
    odoo_user = st.text_input("Usuario", value=st.secrets.get("ODOO_USER", ""), placeholder="admin@empresa.com")
    odoo_pass = st.text_input("Contraseña / API Key", value=st.secrets.get("ODOO_PASSWORD", ""), type="password")
    default_pricelist = st.number_input("ID tarifa por defecto", value=int(st.secrets.get("ODOO_PRICELIST_ID", 2)), min_value=1)
    tag_name = st.text_input("Etiqueta mayorista", value=st.secrets.get("MAYORISTA_TAG_NAME", "Mayorista"))

    st.divider()
    st.subheader("📧 Email (Gmail)")
    smtp_user = st.text_input("Gmail", value=st.secrets.get("SMTP_USER", ""), placeholder="tu@gmail.com")
    smtp_pass = st.text_input("App Password", value=st.secrets.get("SMTP_PASSWORD", ""), type="password")
    from_name = st.text_input("Nombre remitente", value=st.secrets.get("EMAIL_FROM_NAME", "Bio Zen"))

    st.divider()
    st.subheader("🖼️ Logo")
    logo_file = st.file_uploader("Subir logo (PNG)", type=["png", "jpg", "jpeg"])
    logo_bytes = logo_file.read() if logo_file else None

    st.divider()
    listas_sel = st.multiselect(
        "Listas a generar",
        options=list(LISTA_CONFIG.keys()),
        default=list(LISTA_CONFIG.keys()),
        format_func=lambda x: f"{LISTA_CONFIG[x]['icon']} {x}"
    )


# ─── Check connection ───
def try_connect():
    if not odoo_url or not odoo_db or not odoo_user or not odoo_pass:
        return None
    try:
        return get_odoo_connection(odoo_url, odoo_db, odoo_user, odoo_pass)
    except Exception as e:
        return str(e)


# ─── TABS ───
tab_dash, tab_contacts, tab_generate, tab_send = st.tabs([
    "📊 Dashboard", "👥 Contactos", "📁 Generar Listas", "📧 Enviar"
])


# ═══════════════════════════════════════════════════════════════
# TAB 1: DASHBOARD
# ═══════════════════════════════════════════════════════════════
with tab_dash:
    if not odoo_url or not odoo_pass:
        st.info("👈 Configura tus credenciales de Odoo en la barra lateral para comenzar.")
        st.markdown("""
        **Pasos para usar esta herramienta:**

        1. Llena las credenciales de Odoo en la barra lateral
        2. Configura tu App Password de Gmail para los envíos
        3. Sube tu logo de Bio Zen
        4. Ve a la pestaña **Contactos** para verificar tus mayoristas
        5. Ve a **Generar Listas** para crear los XLSX
        6. Ve a **Enviar** para distribuir las listas por email
        """)
    else:
        result = try_connect()
        if isinstance(result, str):
            st.error(f"Error de conexión: {result}")
        elif result:
            odoo = result
            st.success("✓ Conectado a Odoo")

            contacts = get_wholesale_contacts(odoo, tag_name)
            pricelists = get_pricelists(odoo)

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Contactos mayoristas", len(contacts))
            c2.metric("Tarifas activas", len([p for p in pricelists if p['active']]))
            c3.metric("Listas a generar", len(listas_sel))

            by_pl = defaultdict(list)
            for c in contacts:
                by_pl[c['pricelist_name']].append(c)
            c4.metric("Tarifas con clientes", len(by_pl))

            if contacts:
                st.subheader("Distribución por tarifa")
                for pl_name, pl_contacts in sorted(by_pl.items()):
                    with st.expander(f"📋 {pl_name} — {len(pl_contacts)} contactos"):
                        for c in pl_contacts:
                            city = f" ({c['city']})" if c.get('city') else ""
                            st.text(f"  {c['name']}{city} — {c['email']}")


# ═══════════════════════════════════════════════════════════════
# TAB 2: CONTACTS
# ═══════════════════════════════════════════════════════════════
with tab_contacts:
    result = try_connect()
    if isinstance(result, str):
        st.error(f"Error: {result}")
    elif result:
        odoo = result

        if st.button("🔄 Refrescar contactos", type="primary"):
            st.cache_resource.clear()

        contacts = get_wholesale_contacts(odoo, tag_name)

        if not contacts:
            st.warning(f"No se encontraron contactos con la etiqueta '{tag_name}'")
            st.markdown(f"""
            **Para agregar contactos mayoristas:**
            1. Ve a **Contactos** en Odoo
            2. Abre el contacto mayorista
            3. Agrega la etiqueta **"{tag_name}"**
            4. Verifica que tenga asignada la **Lista de precios** correcta
            5. Guarda y refresca aquí
            """)
        else:
            st.success(f"✓ {len(contacts)} contactos mayoristas encontrados")

            # Show as table
            import pandas as pd
            df = pd.DataFrame([
                {
                    'Nombre': c['name'],
                    'Email': c['email'],
                    'Ciudad': c.get('city', ''),
                    'Tarifa': c['pricelist_name'],
                }
                for c in contacts
            ])
            st.dataframe(df, use_container_width=True, hide_index=True)

            # Pricelists
            st.subheader("Tarifas disponibles")
            pricelists = get_pricelists(odoo)
            df_pl = pd.DataFrame([
                {
                    'ID': p['id'],
                    'Nombre': p['name'],
                    'Moneda': p['currency_id'][1] if p['currency_id'] else 'N/A',
                    'Activa': '✓' if p['active'] else '✗',
                }
                for p in pricelists
            ])
            st.dataframe(df_pl, use_container_width=True, hide_index=True)
    else:
        st.info("Configura las credenciales de Odoo en la barra lateral.")


# ═══════════════════════════════════════════════════════════════
# TAB 3: GENERATE LISTS
# ═══════════════════════════════════════════════════════════════
with tab_generate:
    result = try_connect()
    if isinstance(result, str):
        st.error(f"Error: {result}")
    elif result:
        odoo = result

        st.markdown("Selecciona una tarifa y genera las listas con los precios correspondientes.")

        pricelists = get_pricelists(odoo)
        active_pls = [p for p in pricelists if p['active']]
        pl_options = {f"{p['name']} (ID: {p['id']})": p['id'] for p in active_pls}

        selected_pl_label = st.selectbox("Tarifa a usar", options=list(pl_options.keys()))
        selected_pl_id = pl_options[selected_pl_label] if selected_pl_label else default_pricelist
        selected_pl_name = selected_pl_label.split(" (ID:")[0] if selected_pl_label else ""

        if st.button("🚀 Generar listas", type="primary"):
            generated_files = {}
            progress = st.progress(0, text="Iniciando...")

            for i, list_name in enumerate(listas_sel):
                config = LISTA_CONFIG[list_name]
                progress.progress(
                    (i) / len(listas_sel),
                    text=f"Generando {config['icon']} {list_name}..."
                )

                products = get_products_for_list(odoo, config, selected_pl_id)
                if products:
                    xlsx_bytes = generate_xlsx_bytes(
                        list_name, products, config,
                        selected_pl_name, logo_bytes
                    )
                    generated_files[list_name] = {
                        'bytes': xlsx_bytes,
                        'count': len(products),
                        'filename': config['filename'],
                    }

            progress.progress(1.0, text="✓ Listas generadas")

            if generated_files:
                st.session_state['generated_files'] = generated_files
                st.session_state['generated_pl_name'] = selected_pl_name

                st.success(f"✓ {len(generated_files)} listas generadas con tarifa: {selected_pl_name}")

                for name, data in generated_files.items():
                    config = LISTA_CONFIG[name]
                    col1, col2 = st.columns([3, 1])
                    col1.markdown(f"**{config['icon']} {name}** — {data['count']} productos")
                    col2.download_button(
                        "⬇ Descargar",
                        data=data['bytes'],
                        file_name=data['filename'],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_{name}"
                    )
            else:
                st.warning("No se encontraron productos. Verifica las categorías en Odoo.")
    else:
        st.info("Configura las credenciales de Odoo en la barra lateral.")


# ═══════════════════════════════════════════════════════════════
# TAB 4: SEND EMAILS
# ═══════════════════════════════════════════════════════════════
with tab_send:
    result = try_connect()
    if isinstance(result, str):
        st.error(f"Error: {result}")
    elif result:
        odoo = result

        if not smtp_user or not smtp_pass:
            st.warning("Configura las credenciales de Gmail en la barra lateral para enviar emails.")
        else:
            contacts = get_wholesale_contacts(odoo, tag_name)

            if not contacts:
                st.warning("No hay contactos mayoristas. Ve a la pestaña Contactos.")
            else:
                st.subheader("Modo de envío")
                send_mode = st.radio(
                    "¿Cómo quieres enviar?",
                    ["📨 A todos los mayoristas (cada uno con su tarifa)",
                     "📧 A contactos seleccionados",
                     "✉️ A un email específico (prueba)"],
                    horizontal=True
                )

                if "A todos" in send_mode:
                    by_pl = defaultdict(list)
                    for c in contacts:
                        by_pl[c['pricelist_name']].append(c)

                    st.info(f"Se enviarán a **{len(contacts)} contactos** agrupados en **{len(by_pl)} tarifas**")

                    for pl_name, pl_contacts in sorted(by_pl.items()):
                        names = ", ".join([c['name'] for c in pl_contacts[:3]])
                        extra = f" y {len(pl_contacts)-3} más" if len(pl_contacts) > 3 else ""
                        st.markdown(f"- **{pl_name}**: {names}{extra}")

                    if st.button("🚀 Generar y enviar a todos", type="primary"):
                        overall_progress = st.progress(0, text="Iniciando...")
                        status = st.empty()
                        sent_count = 0
                        fail_count = 0
                        total = len(contacts)

                        for pl_name, pl_contacts in by_pl.items():
                            pl_id = pl_contacts[0]['pricelist_id'] or default_pricelist
                            status.info(f"Generando listas para tarifa: {pl_name}...")

                            # Generate files for this pricelist
                            attachment_files = []
                            lists_info = []
                            for list_name in listas_sel:
                                config = LISTA_CONFIG[list_name]
                                products = get_products_for_list(odoo, config, pl_id)
                                if products:
                                    xlsx_bytes = generate_xlsx_bytes(
                                        list_name, products, config, pl_name, logo_bytes
                                    )
                                    attachment_files.append((config['filename'], xlsx_bytes))
                                    lists_info.append({
                                        'name': config['display'],
                                        'icon': config['icon'],
                                        'count': len(products)
                                    })

                            # Send to each contact
                            today = datetime.now().strftime('%d/%m/%Y')
                            subject = f'Bio Zen — Listas de Precios Mayorista ({today})'

                            for contact in pl_contacts:
                                status.info(f"Enviando a {contact['name']}...")
                                try:
                                    html = create_email_html(contact['name'], lists_info, pl_name)
                                    send_email(
                                        'smtp.gmail.com', 587,
                                        smtp_user, smtp_pass, from_name,
                                        contact['email'], contact['name'],
                                        subject, html, attachment_files
                                    )
                                    sent_count += 1
                                except Exception as e:
                                    fail_count += 1
                                    st.warning(f"Error con {contact['email']}: {e}")

                                overall_progress.progress(
                                    (sent_count + fail_count) / total,
                                    text=f"Enviando... {sent_count + fail_count}/{total}"
                                )
                                time.sleep(1)

                        overall_progress.progress(1.0, text="✓ Completado")
                        status.empty()
                        st.success(f"✓ {sent_count} enviados, {fail_count} fallidos de {total} contactos")

                elif "seleccionados" in send_mode:
                    contact_names = [f"{c['name']} ({c['email']}) — {c['pricelist_name']}" for c in contacts]
                    selected = st.multiselect("Selecciona contactos:", contact_names)

                    if selected and st.button("📧 Enviar a seleccionados", type="primary"):
                        sel_contacts = [contacts[contact_names.index(s)] for s in selected]

                        by_pl = defaultdict(list)
                        for c in sel_contacts:
                            by_pl[c['pricelist_name']].append(c)

                        progress = st.progress(0)
                        sent = 0

                        for pl_name, pl_contacts in by_pl.items():
                            pl_id = pl_contacts[0]['pricelist_id'] or default_pricelist

                            attachment_files = []
                            lists_info = []
                            for list_name in listas_sel:
                                config = LISTA_CONFIG[list_name]
                                products = get_products_for_list(odoo, config, pl_id)
                                if products:
                                    xlsx_bytes = generate_xlsx_bytes(list_name, products, config, pl_name, logo_bytes)
                                    attachment_files.append((config['filename'], xlsx_bytes))
                                    lists_info.append({'name': config['display'], 'icon': config['icon'], 'count': len(products)})

                            today = datetime.now().strftime('%d/%m/%Y')
                            subject = f'Bio Zen — Listas de Precios Mayorista ({today})'

                            for contact in pl_contacts:
                                try:
                                    html = create_email_html(contact['name'], lists_info, pl_name)
                                    send_email('smtp.gmail.com', 587, smtp_user, smtp_pass, from_name,
                                               contact['email'], contact['name'], subject, html, attachment_files)
                                    sent += 1
                                except Exception as e:
                                    st.warning(f"Error: {e}")
                                progress.progress(sent / len(sel_contacts))
                                time.sleep(1)

                        st.success(f"✓ {sent} emails enviados")

                elif "específico" in send_mode:
                    test_email = st.text_input("Email de prueba", placeholder="test@email.com")
                    test_pl = st.selectbox(
                        "Tarifa para la prueba",
                        options=[f"{p['name']} (ID: {p['id']})" for p in get_pricelists(odoo) if p['active']],
                        key="test_pl"
                    )

                    if test_email and st.button("✉️ Enviar prueba", type="primary"):
                        test_pl_id = int(test_pl.split("ID: ")[1].rstrip(")"))
                        test_pl_name = test_pl.split(" (ID:")[0]

                        with st.spinner("Generando y enviando..."):
                            attachment_files = []
                            lists_info = []
                            for list_name in listas_sel:
                                config = LISTA_CONFIG[list_name]
                                products = get_products_for_list(odoo, config, test_pl_id)
                                if products:
                                    xlsx_bytes = generate_xlsx_bytes(list_name, products, config, test_pl_name, logo_bytes)
                                    attachment_files.append((config['filename'], xlsx_bytes))
                                    lists_info.append({'name': config['display'], 'icon': config['icon'], 'count': len(products)})

                            today = datetime.now().strftime('%d/%m/%Y')
                            subject = f'Bio Zen — Listas de Precios Mayorista ({today})'
                            html = create_email_html(test_email.split('@')[0], lists_info, test_pl_name)

                            try:
                                send_email('smtp.gmail.com', 587, smtp_user, smtp_pass, from_name,
                                           test_email, test_email, subject, html, attachment_files)
                                st.success(f"✓ Email de prueba enviado a {test_email}")
                            except Exception as e:
                                st.error(f"Error: {e}")
    else:
        st.info("Configura las credenciales de Odoo en la barra lateral.")
